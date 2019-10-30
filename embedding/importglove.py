# 用gensim打开glove词向量需要在向量的开头增加一行：所有的单词数 词向量的维度
import gensim
import os
import shutil
import hashlib
from sys import platform
import numpy as np
import nltk

# 计算行数，就是单词数
from openpyxl import load_workbook

def getFileLineNums(filename):
    f = open(filename, 'r',encoding='utf-8')
    count = 0
    for line in f:
        count += 1
    return count
# Linux或者Windows下打开词向量文件，在开始增加一行
def prepend_line(infile, outfile, line):
    with open(infile, 'r',encoding='utf-8') as old:
        with open(outfile, 'w') as new:
            new.write(str(line) + "\n")
            shutil.copyfileobj(old, new)


def prepend_slow(infile, outfile, line):
    with open(infile, 'r',encoding='utf-8') as fin:
        with open(outfile, 'w',encoding='utf-8') as fout:
            fout.write(line + "\n")
            for line in fin:
                fout.write(line)


def load(filename):
    # num_lines = getFileLineNums(filename)
    gensim_file = 'glove_model.txt'
    # gensim_first_line = "{} {}".format(num_lines, 300)
    # # Prepends the line.
    # if platform == "linux" or platform == "linux2":
    #     prepend_line(filename, gensim_file, gensim_first_line)
    # else:
    #     prepend_slow(filename, gensim_file, gensim_first_line)

    model = gensim.models.KeyedVectors.load_word2vec_format(gensim_file)
    return model

model=load('glove.840B.300d.txt')
wb = load_workbook('worldoil2.xlsx')
from openpyxl import Workbook
wb1 = Workbook()
sheet2 = wb1.active
sheet2.title = 'gloveEach'
from sklearn.feature_extraction.text import TfidfVectorizer
text = []
row = 0
t = 0
for line in open("cleanedNews.txt", "r"):
    text.append(line)  # len = 25589
count = 0
all = []  # 保存所有的新闻向量
print("start calculating...")
for corpus in text:  # 每篇文章
    row += 1
    print("row:", row)
    count += 1
    v = np.zeros(300)
    context = nltk.word_tokenize(corpus)  # 词的list
    corpus1 = [str(corpus)]
    tfidf2 = TfidfVectorizer()
    re = tfidf2.fit_transform(corpus1)
    # 得到一天的向量
    for name in context:  # 每个词
        if name in model:
            j = tfidf2.vocabulary_[name]
            vec = model[str(name)]  # 300维
            vec1 = np.array(vec)
            vec1 = vec1 * re.A[0][j]  # 词的向量乘以tfidf权值
            v+=vec1#把一篇文章的所有词向量加起来
    #v = v * senti  # 乘情感
    all.append(v)
row = 0
print("write excel...")
for day in all:
    row += 1
    for i in range(300):
        sheet2.cell(row, i + 1).value = str(day[i])

wb1.save("glove.xlsx")
