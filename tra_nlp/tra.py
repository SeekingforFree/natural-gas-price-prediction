import openpyxl
import datetime
from openpyxl import load_workbook
import gensim
import numpy as np
from gensim.models.doc2vec import *
import matplotlib.pyplot as plt


#TextBlob,Sheet1,2274,100
def readvec(path,sheet,row,column):
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet)
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name("Sheet1")
    lala=[]
    vec=[]
    y=[]
    num =3
    for i in range(row):
        for j in range(column):
            lala.append(sheet.cell(i+1,j+1).value)
        vec.append(lala)
        y.append(int(sheet1.cell(i+1,num).value))
        # print(sheet1.cell(i+1,4).value)
    print("lr"+str(num))
    return vec,y

#random Forest
def Classifier(train_vecs, y_train, test_vecs, y_test):
    from sklearn.neighbors import KNeighborsClassifier
    from sklearn.ensemble import RandomForestClassifier
    from openpyxl import Workbook

    clf = RandomForestClassifier(n_estimators=5,max_features=0.8,min_samples_leaf=50)
    clf .fit(train_vecs, y_train)
    score = clf.score(test_vecs, y_test)
    print("y_test:")
    print(y_test)
    print('test精度为%s' % score)
    print('train精度为%s' % clf.score(train_vecs, y_train))

def Classifier1(train_list_X,train_list_Y,test_list_X,test_list_Y,x_test,y_test,x_train, y_train):
    from sklearn.svm import SVC
    import random
    X = [[0, 0], [1, 1], [2, 2], [3, 3]]
    Y = [0, 1, 2, 3]
    clf = SVC(probability=True)
    clf.fit(X, Y)
    for epoch in range(10):
        all_reviews_train = train_list_X
        all_reviews_train_Y=train_list_Y
        num = random.randint(0,25536)
        random.seed(num)
        for i in range(len(all_reviews_train)):
            clf.fit(np.array(all_reviews_train[i]),np.array(all_reviews_train_Y[i]))
    score = clf.score(np.array(x_test[0]), np.array(y_test[0]))
    print('test精度为%s' % score)
    print('train精度为%s' % clf.score(np.array(x_train[0]), np.array(y_train[0])))
    return clf

def svm_classifier(train_x, train_y, test_x, test_y):
    import matplotlib.pyplot as plt
    from sklearn import svm
    from sklearn.metrics import roc_curve, auc  ###计算roc和au
    from sklearn.metrics import accuracy_score

    svm = svm.SVC()

    model = svm.fit(train_x, train_y)
    test_y_score = model.decision_function(test_x)
    y_pred = model.predict(test_x)
    predictions = [round(value) for value in y_pred]
    accuracy = accuracy_score(y_test, predictions)
    print("acc:",accuracy)
    print(test_y_score)
    print(predictions)
    # Compute ROC curve and ROC area for each class
    fpr, tpr, threshold = roc_curve(test_y, test_y_score)  ###计算真正率和假正率
    roc_auc = auc(fpr, tpr)  ###计算auc的值

    print("SVM Accuracy:%.2f%%" % (accuracy * 100.0))
    lw = 2
    plt.figure(figsize=(8, 5))
    plt.plot(fpr, tpr, color='darkorange',
             lw=lw, label='ROC curve (area = %0.2f)' % roc_auc)  ###假正率为横坐标，真正率为纵坐标做曲线
    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('svm test roc')
    plt.legend(loc="lower right")
    plt.show()

def navie_classifier(train_x, train_y, test_x, test_y):
    from sklearn.naive_bayes import GaussianNB
    clf = GaussianNB()
    model = clf.fit(train_x, train_y)
    score = clf.score(test_x, test_y)
    print("y_test:")
    print(y_test)
    print('test精度为%s' % score)
    print('train精度为%s' % clf.score(train_x, y_train))

#lightgbm
#xgboost 也有xgboost.train(param)
def xgboost_classifier(x_train,y_train,x_test,y_test):
    #x_train, y_train, x_test, y_test = x_train.tolist(), y_train.tolist(), x_test.tolist(), y_test.tolist()
    from xgboost import XGBClassifier
    from sklearn.metrics import accuracy_score
    for i in range(len(x_train)):
        if(y_train[i]==-1):
            y_train[i]=0
    for i in range(len(y_test)):
        if(y_test[i]==-1):
            y_test[i]=0
    model=XGBClassifier()
    eval_set=[(x_test,y_test)]
    model.fit(x_train,y_train,eval_metric="error",eval_set=eval_set,verbose=True)
    y_pred=model.predict(x_test)
    predictions=[round(value) for value in y_pred]
    accuracy=accuracy_score(y_test,predictions)
    print("Accuracy:%.2f%%"%(accuracy*100.0))

# logical regression
def lr_classifier(x_train,y_train,x_test,y_test):
    from sklearn import linear_model
    lr=linear_model.LogisticRegression()
    lr.fit(x_train,y_train)
    lr.score(x_test,y_test)
    print(lr.predict(x_test))
    print("lr test:",lr.score(x_test,y_test))
    print("lr train:",lr.score(x_train,y_train))
    return lr


vec ,y =readvec('w2v.xlsx','Sheet3',2269,300 )
#s输入为未分类的vec，将vec分为trainvec,testvec,unsvec
def divide(xx,yy):
    return [xx[i] for i in range(0, int(0.6 * len(xx)))], [xx[i] for i in range(int(0.6 * len(xx)), int(0.8 * len(xx)))], [ xx[i] for i in range(int(0.8 * len(xx)), len(xx))],[yy[i] for i in range(0,int(0.6*len(yy)))],[yy[i] for i in range(int(0.6*len(yy)),int(0.8*len(yy)))]


x_train, x_test, unsup_reviews,y_train,y_test=divide(vec,y)
print(len(x_train),len(x_test))
print(len(y_train),len(y_test))

x_train, x_test, unsup_reviews=np.array(x_train,dtype='float16'),np.array( x_test,dtype='float16'), np.array(unsup_reviews,dtype='float16')
y_train, y_test =np.array(y_train),np.array( y_test)
# lr_classifier(x_train,y_train,x_test,y_test)
# xgboost_classifier(x_train,y_train,x_test,y_test)
# svm_classifier(x_train, y_train,x_test, y_test )
navie_classifier(x_train, y_train,x_test, y_test )

def est(y_train,y_test):
    train_zheng=0
    train_fu =0
    test_zheng =0
    test_fu =0
    for i in range(len(y_train)):
        if(y_train[i]==1):
            train_zheng=train_zheng+1
        else:
            train_fu =train_fu+1
    for i in range(len(y_test)):
        if (y_train[i] == 1):
            test_zheng = test_zheng + 1
        else:
            test_fu = test_fu + 1
    print(train_zheng,train_fu,test_zheng,test_fu)
    return train_zheng,train_fu,test_zheng,test_fu


def ROC_curve(lr, y_train):
    from sklearn.metrics import roc_curve, auc
    import matplotlib.pyplot as plt
    print(lr.predict_proba(x_train))
    pred_probas = lr.predict_proba(x_train)[:, 1]

    fpr, tpr, _ = roc_curve(y_train ,pred_probas)
    roc_auc = auc(fpr, tpr)
    plt.plot(fpr, tpr, label='train area = %.2f' % roc_auc)
    plt.plot([0, 1], [0, 1], 'k--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.title('train')
    plt.show()
    return



def ROC_curve1(lr, y_test):
    from sklearn.metrics import roc_curve, auc
    import matplotlib.pyplot as plt

    pred_probas = lr.predict_proba(x_test)[:, 1]

    fpr, tpr, _ = roc_curve(y_test ,pred_probas)
    roc_auc = auc(fpr, tpr)
    plt.plot(fpr, tpr, label='test area = %.2f' % roc_auc)
    plt.plot([0, 1], [0, 1], 'k--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.title('test')
    plt.show()
    return

# ROC_curve1(lr, y_test)