from sklearn.manifold import TSNE
import matplotlib.pyplot as plt
import datetime
from openpyxl import load_workbook
import gensim
import numpy as np
from gensim.models.doc2vec import *


def read():
    wb = load_workbook('w2v.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    wb1 = load_workbook('b.xlsx')
    sheet1 = wb1.get_sheet_by_name("Sheet1")
    vec=[]
    y=[]
    lala=[]
    for i in range(2274):#2275,6都不行
        for j in range(100):
            lala.append(sheet.cell(i+1,j+1).value)
        vec.append(lala)
        lala=[]
        y.append(sheet1.cell(i+1,5).value)
    return  vec,y

vec,y=read()
def tsne():
    ts = TSNE(2)
    reduced_vecs = ts.fit_transform(vec)
    #color points by word group to see if Word2Vec can separate them
    for i in range(len(reduced_vecs)):
        if y[i]==1:
            #zheng words colored blue
            color = 'r'
        elif y[i]==0:
            #bubian words colored red
            color = 'y'
        elif y[i]==-1:
            #weather words colored green
            color = 'g'
        print(reduced_vecs[i,0],",", reduced_vecs[i,1],i,":",color)
        plt.plot(reduced_vecs[i,0], reduced_vecs[i,1], marker='.', color=color, markersize=8)
        plt.title("cleanedNews tsne:r y g")
    plt.show()
#b
# tsne()
def pca():
    from sklearn.decomposition import PCA
    pca=PCA(2)
    reduced_vecs = pca.fit_transform(vec)
    # color points by word group to see if Word2Vec can separate them
    for i in range(len(reduced_vecs)):
        if y[i] == 1:
            # zheng words colored blue
            color = 'r'
        elif y[i] == 0:
            # bubian words colored red
            color = 'y'
        elif y[i] == -1:
            # weather words colored green
            color = 'g'
        print(reduced_vecs[i, 0], ",", reduced_vecs[i, 1], i, ":", color)
        plt.plot(reduced_vecs[i, 0], reduced_vecs[i, 1], marker='.', color=color, markersize=8)
        plt.title("pca:r y g")
    plt.show()

pca()
def u_map():
    import umap
    reducer= umap.UMAP()
    reducer.fit(vec)
    reduced_vecs = reducer.transform(vec)
    for i in range(len(reduced_vecs)):
        if y[i] == 1:
            # zheng words colored blue
            color = 'b'
        elif y[i] == 0:
            # bubian words colored red
            color = 'r'
        elif y[i] == -1:
            # weather words colored green
            color = 'y'
        print(reduced_vecs[i, 0], ",", reduced_vecs[i, 1], i, ":", color)
        plt.plot(reduced_vecs[i, 0], reduced_vecs[i, 1], marker='.', color=color, markersize=8)
        plt.title('UMAP projection of the Iris dataset', fontsize=24)
    plt.show()

# u_map()