# coding=utf-8

from sklearn.metrics import accuracy_score
from sklearn.model_selection import GridSearchCV
import lightgbm as lgb
# import xgboost as lgb
import numpy as np
import pandas as pd
from openpyxl import load_workbook

def read():#TextBlob.xlsx num_model.xlsx,cleanedNew.xlsx

    wb = load_workbook("finalworldoil2.xlsx")
    sheet = wb.get_sheet_by_name('textblob100')
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name("Sheet1")
    vec = []
    y = []
    lala = []
    for i in range(2274):  # 2275,6都不行
        for j in range(100):
            lala.append(sheet.cell(i + 1, j + 1).value)
        vec.append(lala)
        lala = []
        num=4
        # y.append(sheet1.cell(i + 1,4).value+1)
        if(sheet1.cell(i + 1, num).value==-1):
            y.append(1)
        elif sheet1.cell(i + 1, num).value==0:
            y.append(2)
        elif sheet1.cell(i + 1, num).value == 1:
            y.append(3)
    # fp.write(name+" sheet1."+str(num)+"\n\n")
    return vec, y


def divide(xx, yy):
    return [xx[i] for i in range(0, int(0.6 * len(xx)))], [xx[i] for i in
                                                           range(int(0.6 * len(xx)), int(0.8 * len(xx)))], [xx[i] for i
                                                                                                            in range(
            int(0.8 * len(xx)), len(xx))], [yy[i] for i in range(0, int(0.6 * len(yy)))], [yy[i] for i in
                                                                                           range(int(0.6 * len(yy)),
                                                                                                 int(0.8 * len(yy)))]

vec, y = read()

train_x, test_x, unsup_reviews, train_y, test_y = divide(vec, y)
train_x, test_x, unsup_reviews = np.array(train_x), np.array(test_x), np.array(unsup_reviews)
train_y, test_y = np.array(train_y), np.array(test_y)

train_data = lgb.Dataset(train_x, label=train_y)
test_data = lgb.Dataset(test_x, label=test_y, reference=train_data)

# binary binary_error
param = {
    'task': 'train',
    'boosting_type': 'gbdt',
    'num_leaves': 16,
    'num_trees': 100,
    'objective': 'multiclass',#'binary',
    'metric': 'multi_error',
    'num_class': 3,
    'max_bin': 255,
    'learning_rate': 0.05,
    'early_stopping': 500
}

print("multierror")
#
num_round = 500

bst = lgb.train(param, train_data, num_round, valid_sets=[test_data])
bst.save_model('week_lightgbm_binary2.txt')
mybst = lgb.Booster(model_file='model.txt')  # init model
ypred_train = bst.predict(train_x)
ypred_test = bst.predict(test_x)
print("1",ypred_train)
print(ypred_test)

ypred_train = [list(x).index(max(x)) for x in ypred_train]
ypred_test = [list(x).index(max(x)) for x in ypred_test]



print('The train error rate of prediction is:'+
     str( accuracy_score(train_y, ypred_train))+'\n')
print('The test error rate of prediction is:'+
      str(accuracy_score(test_y, ypred_test))+'\n')

cv_results = lgb.cv(param, train_data, num_round, nfold=5)

print('best n_estimators:'+str( len(cv_results['multi_error-mean']))+'\n')
print('best cv score:'+ str(pd.Series(cv_results['multi_error-mean']).min())+'\n')
