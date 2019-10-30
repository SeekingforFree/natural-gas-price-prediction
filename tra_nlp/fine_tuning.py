# coding=utf-8

from sklearn.metrics import accuracy_score
from sklearn.model_selection import GridSearchCV
import lightgbm as lgb
# import xgboost as lgb
import numpy as np
import pandas as pd
from openpyxl import load_workbook

def read(name,num):#TextBlob.xlsx num_model.xlsx,cleanedNew.xlsx
    # fp =open("train_model.txt","a+")
    wb = load_workbook(name)
    sheet = wb.get_sheet_by_name('Sheet3')
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name("Sheet1")
    vec = []
    y = []
    lala = []
    for i in range(2274):  # 2275,6都不行
        for j in range(300):
            lala.append(sheet.cell(i + 1, j + 1).value)
        vec.append(lala)
        lala = []
        # y.append(sheet1.cell(i + 1, num).value+1)
        if(sheet1.cell(i + 1, num).value==-1):
            y.append(0)
        elif sheet1.cell(i + 1, num).value==1:
            y.append(1)
    print(name+" sheet1."+str(num)+"\n\n")
    return vec, y


def divide(xx, yy):
    return [xx[i] for i in range(0, int(0.6 * len(xx)))], [xx[i] for i in
                                                           range(int(0.6 * len(xx)), int(0.8 * len(xx)))], [xx[i] for i
                                                                                                            in range(
            int(0.8 * len(xx)), len(xx))], [yy[i] for i in range(0, int(0.6 * len(yy)))], [yy[i] for i in
                                                                                           range(int(0.6 * len(yy)),
                                                                                                 int(0.8 * len(yy)))]

# vec, y = read()

def process(vec,y):
    train_x, test_x, unsup_reviews, train_y, test_y = divide(vec, y)
    train_x, test_x, unsup_reviews = np.array(train_x,dtype='float16'), np.array(test_x,dtype='float16'), np.array(unsup_reviews,dtype='float16')
    train_y, test_y = np.array(train_y), np.array(test_y)

    train_data = lgb.Dataset(train_x, label=train_y)
    test_data = lgb.Dataset(test_x, label=test_y, reference=train_data)
    return train_x,test_x,train_y,test_y,train_data,test_data

def model(train_data,test_data,train_x,test_x,train_y,test_y,objective,metric):
    fp =open("train_model.txt","a+")
    param = {
        'task': 'train',
        'boosting_type': 'gbdt',
        'num_leaves': 16,
        'num_trees': 100,
        'objective': 'multiclassova',
        'metric': 'multi_error',
        'max_bin': 255,
        'learning_rate': 0.05,
        'early_stopping': 10
    }
    param.update({'objective':objective})
    param.update({'metric': metric})
    print(param)
    print("multierror")

    num_round = 10

    bst = lgb.train(param, train_data, num_round, valid_sets=[test_data])
    bst.save_model('model.txt')
    mybst = lgb.Booster(model_file='model.txt')  # init model
    ypred_train = bst.predict(train_x)
    ypred_test = bst.predict(test_x)
    print("ypred_train",ypred_train)
    print("ypred_test", ypred_test)
    train_pred=[]
    for i in range(len(ypred_train)):
        if ypred_train[i]>0.5:
            train_pred.append(1)
        else:
            train_pred.append(0)
    test_pred = []
    for i in range(len(ypred_test)):
        if ypred_test[i] > 0.5:
            test_pred.append(1)
        else:
            test_pred.append(0)

    print('The train error rate of prediction is:'
         ,( accuracy_score(train_y, train_pred)))
    print('The test error rate of prediction is:',
          (accuracy_score(test_y, test_pred)))

    cv_results = lgb.cv(param, train_data, num_round, nfold=5)
    name =metric+"-mean"
    print('best n_estimators:',( len(cv_results[name])))
    print('best cv score:',(pd.Series(cv_results[name]).min()),'\n')

def readDouble():
    vec_train_list=[]
    y_list=[]
    vec, y=read("num_model.xlsx",1)
    vec_train_list.append(vec)
    y_list.append(y)

    return vec_train_list,y_list

objective_list=["fair","binary"]
metric_list=["fair","binary_error","binary_logloss","auc"]# auc
print("add auc")
vec,y =readDouble()
for i in range(len(vec)):
    train_x, test_x, train_y, test_y, train_data, test_data=process(vec[i],y[i])
    for metric in metric_list:
        try:
            model(train_data,test_data,train_x,test_x,train_y,test_y,objective_list[0],metric)
        except AssertionError:
            print("AssertionError")
        else:
            print(objective_list[0],metric)
