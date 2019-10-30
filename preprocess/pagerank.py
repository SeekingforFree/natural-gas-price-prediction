import numpy as np
import pandas as pd
import nltk
import networks as nx

word_embedding={}
f= open('glove.6B.100d.txt',encoding='utf-8')
for line in f :
    values =line.strip().split()
    word =values[0]
    coefs =np.asarray(values[1:],dtype='float32')
    word_embedding[word]=coefs
f.close()

def readfile():
    from openpyxl import load_workbook
    from nltk.tokenize import sent_tokenize
    wb = load_workbook('wor.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    sentences = []
    for i in range(25599):
        sentences.append(sent_tokenize(sheet.cell(i + 1, 3).value))
    sentences = [y for x in sentences for y in x]  # flatten list
    print(sentences[:10])
    print("sentences length:",len(sentences))
    return sentences

sentence=readfile()




#文本预处理
clean_sentence=pd.Series(sentence).str.replace('[.]',"")
#生成句子的特征向量
sentence_vec=[]
for i in clean_sentence:
    if len(i)!= 0:
        v= sum([word_embedding.get(w,np.zeros((100,))) for w in i.split()])/(len(i.split())+0.001)
    else:
        v=np.zeros((100,))
    sentence_vec.append(v)

sim_mat=np.zeros([len(sentence),len(sentence)])
from sklearn.metrics.pairwise import cosine_similarity
for i in range(len(sentence)):
    for j in range(len(sentence)):
        if i!=j:
            sim_mat[i][j] = cosine_similarity(sentence_vec[i].reshape(1,100),sentence_vec[j].reshape(1,100))[0,0]


nx_graph =nx.from_numpy_array(sim_mat)
scores=nx.pagerank(nx_graph)

ranked_sentence=sorted(((scores[i],s)for i,s in enumerate(clean_sentence)),reverse=True)

for i in range(10):
    print(ranked_sentence[i])
