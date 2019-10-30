from openpyxl import load_workbook
from nltk.tokenize import sent_tokenize
wb = load_workbook('wor.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
fp = open('cleanedNews.txt', "w", encoding=u'utf-8', errors='ignore')

for i in range(25599):
    line = sheet.cell(i+2,3).value
    import re

    # temp = f.read()
    # temp = temp.decode("utf8")
    string = re.sub(r'[^a-zA-Z. ]', "", line)
    string = re.sub(r' [a-z] ', "", string)
    if len(string.split(' ')[0])==1:
        fp.write(string.split(' ',1)[1].lower()+'\n')
    else:
        fp.write(string.lower()+'\n')

fp.close()