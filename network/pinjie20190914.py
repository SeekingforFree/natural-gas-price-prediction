import keras
from keras import Model, Sequential
from keras.layers import Input, Dense, merge, LSTM, Dropout, Conv1D, MaxPooling1D, Bidirectional, Multiply, Conv2D, \
    MaxPooling2D, TimeDistributed, Reshape, Flatten, Concatenate
import numpy as np
from openpyxl import load_workbook


def readvec(path,sheetname,length=300,num=2,column=25589):
    # 读情感值
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name(sheetname)
    #读价格
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    # 读日期标签
    wb2 =load_workbook('glove.xlsx')# worldoil2.xlsx')情感值
    sheet2 =wb2.get_sheet_by_name('glove')
    lala=[]
    # 每天的价格
    vec=[]
    # 总的返回向量
    vector=[]
    y=[]
    date =sheet2.cell(1,1).value

    for i in range(column):
        new_date =sheet2.cell(i+1,1).value
        if new_date!=date:
            print(date,new_date)
            date=new_date
            vector.append(vec)
            vec=[]
            lala=[]
        # print(date,new_date)
        for j in range(length):
            lala.append(float(sheet.cell(i+1,j+1).value))
        vec.append(lala)
        lala=[]
    # vector.append(vec)
    if num ==4:
        for i in range(2274):
            # #row =5 三分类
            if sheet1.cell(i+1,4).value==-1:
                y.append(0)
            elif sheet1.cell(i+1,4).value==0:
                y.append(1)
            else:
                y.append(2)
    else:
        for i in range(2274):
        #二分类
            if sheet1.cell(i + 1, num).value == -1:
                y.append(0)
            else:
                y.append(1)
        #y.append(int(sheet1.cell(i+1,8).value))
        # print(sheet1.cell(i+1,4).value)
    return vector,y

def reverse(vec):
    length =len(vec)
    vector=[]
    for i in range(length):
        vector.append(vec[length-1-i])
    return vector

def toTimeDistribute(vec,y,l):
    vec=reverse(vec)
    y =reverse(y)
    lala=[]
    vector=[]
    Y=[]
    for i in range(len(vec)-l+1):###+1
        for j in range(l):
            lala.append(vec[i+j])
        vector.append(lala)
        Y.append(y[i+l-1])####-1
        lala=[]
    return vector,Y

class AccuracyHistory(keras.callbacks.Callback):
    def on_train_begin(self, logs={}):
        self.acc = []
        self.loss=[]
        self.val_loss=[]
        self.val_acc=[]

    def on_epoch_end(self, batch, logs={}):
        self.acc.append(logs.get('acc'))
        self.loss.append(logs.get('loss'))
        self.val_acc.append(logs.get('val_acc'))
        self.val_loss.append(logs.get('val_loss'))
def preprocessCNN(filename,sheet,length,num,Timelen,num_classes,column):
    # num=4 三分类num=1,2,3
    vec, y = readvec(filename, sheet, length, num,column=column)
    # import matplotlib.pyplot as plt
    from sklearn.preprocessing import StandardScaler
    X_scale = StandardScaler()
    # vec = X_scale.fit_transform(vec)

    X = []
    lala = []
    for j in range(length): lala.append(0)
    for i in range(len(vec)):
        while len(vec[i]) < 35:
            vec[i].append(lala)
    vec, y = toTimeDistribute(vec, y, Timelen)
    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(vec, y, test_size=0.4,shuffle=False)
    img_x, img_y = 35, length
    x_train=np.array(x_train)
    x_test=np.array(x_test)
    x_train = x_train.reshape(int(x_train.shape[0]), Timelen, img_x, img_y)
    x_test = x_test.reshape(int(x_test.shape[0]), Timelen, img_x, img_y)
    y_train = keras.utils.to_categorical(y_train, num_classes)
    y_test = keras.utils.to_categorical(y_test, num_classes)
    return x_train, y_train, x_test, y_test

# num是以哪种形式读取数据 ：1，0变成1；2，0变成-1；3，0变成与上一天相同；4，输出结果就是-1，0，1
#filenum 从那个文件里读取向量信息 sheet 是xlsx文件中的sheet名
# epochs是神经网络的迭代次数，Timelen是lstm中的持续时间 column=2274
def processCNN(num=2,filename='glove.xlsx',sheet ='glove',length=300, Timelen = 5,column=25589):
    if num == 4:
        num_classes = 3
    else:
        num_classes = 2
    input_shape = (Timelen,35)#, length)#Timelen,
    x_train, y_train,x_test, y_test=preprocessCNN(filename,sheet,length,num,Timelen,num_classes,column=column)
    return input_shape,num_classes,x_train, y_train,x_test, y_test

def toTimeDistributeTime(vec,y,l):
    vec=reverse(vec)
    y =reverse(y)
    lala=[]
    vector=[]
    Y=[]
    for i in range(len(vec)-l):
        for j in range(l):
            lala.append(vec[i+j])
        vector.append(lala)
        Y.append(y[i+l])
        lala=[]
    return vector,Y

def preprocessTime(num,Timelen,num_classes,type=1):
    # num=4 三分类num=1,2,3
    # type=1,x是数值，y是0，1；type=0，x
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    y=[]
    vec=[]

    if num ==4:
        for i in range(2274):
            # #row =5 三分类
            if sheet1.cell(i+1,4).value==-1:
                y.append(0)
            elif sheet1.cell(i+1,4).value==0:
                y.append(1)
            else:
                y.append(2)
    else:
        for i in range(2274):
        #二分类
            if sheet1.cell(i + 1, num).value == -1:
                y.append(0)
            else:
                y.append(1)

    if type==1:
        for i in range(2274):
            vec.append(sheet1.cell(i+1,5).value)
    else:
        vec=y

    vec, y = toTimeDistribute(vec, y, Timelen)
    # return vec ,y
    # import matplotlib.pyplot as plt
    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(vec, y, test_size=0.4,shuffle=False)
    x_train=np.array(x_train)
    x_test=np.array(x_test)
    x_train = x_train.reshape(int(x_train.shape[0]), Timelen,1)
    x_test = x_test.reshape(int(x_test.shape[0]), Timelen,1)
    print('x_train shape:', x_train.shape)
    print(x_train.shape[0], 'train samples')
    print(x_test.shape[0], 'test samples')

    # convert class vectors to binary class matrices - this is for use in the
    # categorical_crossentropy loss below

    y_train = keras.utils.to_categorical(y_train, num_classes)
    y_test = keras.utils.to_categorical(y_test, num_classes)
    return x_train, y_train, x_test, y_test

def processTime(num=2, Timelen = 5):
    if num == 4:
        num_classes = 3
    else:
        num_classes = 2
    # x_train, y_train,x_test, y_test=preprocessTime(num,Timelen,num_classes)
    return preprocessTime(num,Timelen,num_classes)
# merge
def build_model(timestep=5,input_shape=(),num_classes=3):
    input1 = Input(shape=(timestep,1))
    lstm1 = LSTM(64)(input1)
    d1=Dropout(0.4)(lstm1)
    attention_probs1 = Dense(128)(d1)
    attention_probs1 = Dropout(0.5)(attention_probs1)

    input = Input((timestep,35,300))
    td1=TimeDistributed(Conv1D(100, kernel_size=2, strides=1),
                     input_shape=input_shape)(input)
    x=TimeDistributed(MaxPooling1D(pool_size=2, strides=2))(td1)
    kernellist = [3, 4]
    for list in kernellist:
        x=TimeDistributed(Conv1D(100, list))(x)
        x = TimeDistributed(MaxPooling1D(pool_size=2))(x)
    x = TimeDistributed(Flatten())(x)
    x= Dropout(0.5)(x)
    lstm_out = LSTM(64)(x)

    attention_probs = Dense(128)(lstm_out)
    attention_mul = Concatenate()([attention_probs,attention_probs1])
    attention_mul = Dropout(0.5)(attention_mul)
    output = Dense(num_classes, activation='softmax')(attention_mul)
    model = Model(inputs=[input,input1], outputs=output)
    print(model.summary())


    model.compile(loss=keras.losses.categorical_crossentropy,
                  optimizer=keras.optimizers.Adam(),
                  metrics=['accuracy'])
    return model


def process2(num=2,batch_size=128, \
    epochs = 500, Timelen = 5):
    if num == 4:
        num_classes = 3
    else:
        num_classes = 2


    x_train, y_train,x_test, y_test=preprocessTime(num,Timelen,num_classes)
    history = AccuracyHistory()
    model=Sequential()
    model.add(LSTM(Timelen))
    model.add(Dropout(0.4))
    model.add(Dense(num_classes, activation='softmax'))
    model.compile(loss=keras.losses.categorical_crossentropy,
                  optimizer=keras.optimizers.Adam(),
                  metrics=['accuracy'])
    model.fit(x_train, y_train,
              batch_size=batch_size,
              epochs=epochs,
              verbose=1,
              validation_data=(x_test, y_test),
              callbacks=[history])
    # joblib.dump(clf,"train_model.m") 保存模型
    # clf2 =joblib.load("train_model.m")
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])

    # # 加载整个模型
    # model.load_model('my_model.h5')
    return history.acc,history.loss,history.val_acc,history.val_loss

# length是向量的维度
def process(filename='glove.xlsx',sheet='glove',batch_size=128, epochs = 30,Timelen=5,length=300,num=2):
    input_shape,num_classes,x_train1, y_train1, x_test1, y_test1=processCNN(filename=filename,Timelen=Timelen,sheet=sheet,num=num,length=length)
    model = build_model(timestep=Timelen,input_shape=input_shape, num_classes=num_classes)
    x_train2, y_train2, x_test2, y_test2=processTime(num=num,Timelen=Timelen)
    print(y_train1)
    print(y_train2)
    # for i in range(1362):
    #     print(y_train1[i],",",y_train2[i])
    history = AccuracyHistory()

    x_train=[x_train1,x_train2]
    x_test =[x_test1,x_test2]
    model.fit(x_train, y_train1,
              batch_size=batch_size,
              epochs=epochs,
              verbose=1,
              validation_data=(x_test, y_test1),
              callbacks=[history])
    # joblib.dump(clf,"train_model.m") 保存模型
    # clf2 =joblib.load("train_model.m")
    score = model.evaluate(x_test, y_test1, verbose=0)
    ypred_train = model.predict(x_train)
    ypred_test = model.predict(x_test)

    print("train pred:",ypred_train)
    print("test pred:", ypred_test)
    import matplotlib.pylab as plt
    plt.plot(range(1, epochs + 1), history.acc)
    plt.xlabel('Epochs')
    plt.ylabel(sheet + 'Accuracy with Dropout' +str(num))
    plt.show()

    plt.plot(range(1, epochs + 1), history.loss)
    plt.xlabel('Epochs')
    plt.ylabel(sheet + 'Loss with Dropout' + str(num))
    plt.show()
    return history.acc,history.loss,history.val_acc,history.val_loss

def __main__():
    from openpyxl import Workbook

    count =1
    filename='finalworldoil2.xlsx'
    # sheet1=['nosenti100','nltk100','textblob100']# 0 2274 0 2274 0 2274
    sheet300=['w2vEach']
    sheet100=['w2v','w2v2','w2vNosenti']
    times=[7,30,60,120,300]
    nums=[1,2,3,4]
    for time in times:
        for sheet in sheet300:
            for num in nums:
                # acc, loss, testacc, testloss =process2(num=num)
                if count>12:
                    wb = Workbook()

                    sheet1 = wb.active
                    sheet1.title = 'acc_and_loss'
                    acc, loss, testacc, testloss =process(filename=filename,Timelen=time,sheet=sheet,epochs=200,num=num)
                    raw =1
                    sheet1.cell(raw, 1).value = str(filename) + " " + str(sheet) + " " + str(num) + " " + str(time)
                    raw += 1
                    for i in range(len(testacc)):
                        sheet1.cell(raw, i + 1).value = testacc[i]
                    raw += 1
                    for i in range(len(testloss)):
                        sheet1.cell(raw, i + 1).value = testloss[i]
                    raw += 1
                    for i in range(len(acc)):
                        sheet1.cell(raw, i + 1).value = acc[i]
                    raw += 1
                    for i in range(len(loss)):
                        sheet1.cell(raw, i + 1).value = loss[i]
                    raw += 1
                    wb.save("model/finalweek_acc_and_loss_nonsenti"+str(count)+".xlsx")
                count = count+1
__main__()