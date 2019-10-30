import keras
from keras.layers import Dense, Flatten, LSTM, MaxPooling1D, Convolution1D, Conv1D, \
    Dropout, TimeDistributed, MaxPooling2D, Conv2D
from keras.models import Sequential
import matplotlib.pylab as plt

from openpyxl import load_workbook
import numpy as np

def readvec(path,sheet,length=300,num=4,column=25589):
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet)
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    wb2 =load_workbook('worldoil2.xlsx')
    sheet2 =wb2.get_sheet_by_name('Sheet2')
    # wb3 = load_workbook('w2veach.xlsx')  # worldoil2.xlsx')情感值
    # sheet3 = wb3.get_sheet_by_name('w2v')
    lala=[]
    vec=[]
    vector=[]
    y=[]
    date =sheet.cell(1,1).value

    for i in range(column):
        new_date =sheet.cell(i+1,1).value
        if new_date!=date:
            print(date,new_date)
            date=new_date
            vector.append(vec)
            vec=[]
            lala=[]
        # print(date,new_date)
        for j in range(length):
            lala.append(float(float(sheet2.cell(i+11,4).value)*sheet.cell(i+1,j+2).value))
        vec.append(lala)
        lala=[]
    # vector.append(vec)
    if num ==4:
        for i in range(2274):
            # #row =5 三分类
            if sheet1.cell(i+1,4).value==-1:
                y.append(0)
            elif sheet1.cell(i+1,4).value==0:
                y.append(1)
            else:
                y.append(2)
    else:
        for i in range(2274):
        #二分类
            if sheet1.cell(i + 1, num).value == -1:
                y.append(0)
            else:
                y.append(1)
        #y.append(int(sheet1.cell(i+1,8).value))
        # print(sheet1.cell(i+1,4).value)
    return vector,y

def reverse(vec):
    length =len(vec)
    vector=[]
    for i in range(length):
        vector.append(vec[length-1-i])
    return vector

def toTimeDistribute(vec,y,l):
    vec=reverse(vec)
    y =reverse(y)
    lala=[]
    vector=[]
    Y=[]
    for i in range(len(vec)-l+1):
        for j in range(l):
            lala.append(vec[i+j])
        vector.append(lala)
        Y.append(y[i+l-1])
        lala=[]
    return vector,Y

# num_classes is output dim

def build_model(input_shape,num_classes):
    model = Sequential()
    model.add(TimeDistributed(Conv1D(100, kernel_size=2, strides=1),
                     input_shape=input_shape))
    model.add(TimeDistributed(MaxPooling1D(pool_size=2, strides=2)))
    kernellist=[3,4]
    for list in kernellist:
        model.add(TimeDistributed(Conv1D(100, list)))
        model.add(TimeDistributed(MaxPooling1D(pool_size=2)))
    model.add(TimeDistributed(Dropout(0.4)))
    model.add(TimeDistributed(Flatten()))

    model.add(LSTM(50,dropout=0.4,))#return_sequence
    model.add(Dense(1000))

    model.add(Dense(num_classes, activation='softmax'))

    model.compile(loss=keras.losses.categorical_crossentropy,
                  optimizer=keras.optimizers.Adam(),
                  metrics=['accuracy'])
    return model

def build_model2D(input_shape,num_classes):
    model = Sequential()
    model.add(Conv2D(100, kernel_size=2, strides=1))
    model.add(MaxPooling2D(pool_size=2, strides=2))
    model.add(Dropout(0.4))
    # model.add(ZeroPadding2D(padding=(1, 1),data_format='channels_first'))
    model.add(Flatten())

    model.add(Dense(1000))

    model.add(Dense(num_classes, activation='softmax'))

    model.compile(loss=keras.losses.categorical_crossentropy,
                  optimizer=keras.optimizers.Adam(),
                  metrics=['accuracy'])
    return model


class AccuracyHistory(keras.callbacks.Callback):
    def on_train_begin(self, logs={}):
        self.acc = []
        self.loss=[]
        self.val_loss=[]
        self.val_acc=[]

    def on_epoch_end(self, batch, logs={}):
        self.acc.append(logs.get('acc'))
        self.loss.append(logs.get('loss'))
        self.val_acc.append(logs.get('val_acc'))
        self.val_loss.append(logs.get('val_loss'))

def preprocess(filename,sheet,length,num,Timelen,num_classes,column):
    # num=4 三分类num=1,2,3
    vec, y = readvec(filename, sheet, length, num,column=column)
    # import matplotlib.pyplot as plt
    from sklearn.preprocessing import StandardScaler
    X_scale = StandardScaler()
    # vec = X_scale.fit_transform(vec)

    X = []
    lala = []
    for j in range(length): lala.append(0)
    for i in range(len(vec)):
        while len(vec[i]) < 35:
            vec[i].append(lala)
        # X.append(X_scale.fit_transform((np.array(vec[i],dtype='float16').reshape(35*200,1))))
    vec, y = toTimeDistribute(vec, y, Timelen)
    X = np.array(vec)
    # X=X.reshape(2274,35*200,1)
    print(type(X))
    print(X.shape)
    # print(X)
    y = np.array(y)

    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=0.4)
    # input image dimensions
    img_x, img_y = 35, length

    print(x_train.shape)
    print(type(x_train))
    # print(x_train)
    # reshape the data into a 4D tensor - (sample_number, x_img_size, y_img_size, num_channels)
    # because the MNIST is greyscale, we only have a single channel - RGB colour images would have 3
    x_train = x_train.reshape(int(x_train.shape[0]), Timelen, img_x, img_y)
    x_test = x_test.reshape(int(x_test.shape[0]), Timelen, img_x, img_y)
    print('x_train shape:', x_train.shape)
    print(x_train.shape[0], 'train samples')
    print(x_test.shape[0], 'test samples')

    # convert class vectors to binary class matrices - this is for use in the
    # categorical_crossentropy loss below

    y_train = keras.utils.to_categorical(y_train, num_classes)
    y_test = keras.utils.to_categorical(y_test, num_classes)
    return x_train, y_train, x_test, y_test

# num是以哪种形式读取数据 ：1，0变成1；2，0变成-1；3，0变成与上一天相同；4，输出结果就是-1，0，1
#filenum 从那个文件里读取向量信息 sheet 是xlsx文件中的sheet名
# epochs是神经网络的迭代次数，Timelen是lstm中的持续时间
def process(num=2,filename='glove.xlsx',sheet ='gloveEach',length=300,batch_size=128, \
    epochs = 300, Timelen = 5,column=25589):
    if num == 4:
        num_classes = 3
    else:
        num_classes = 2


    input_shape = (Timelen, 35, length)

    x_train, y_train,x_test, y_test=preprocess(filename,sheet,length,num,Timelen,num_classes,column=column)
    history = AccuracyHistory()
    model = build_model(input_shape,num_classes)
    model.fit(x_train, y_train,
              batch_size=batch_size,
              epochs=epochs,
              verbose=1,
              validation_data=(x_test, y_test),
              callbacks=[history])
    score = model.evaluate(x_test, y_test, verbose=0)
    print("y train:",y_train)
    k=0
    for i in range(len(y_train)):
        if y_train[i][0]==1:
            k=k+1
    print(k)
    print("x train:",x_train)
    y_pred = model.predict(x_train)
    ypred_train =[list(x).index(max(x)) for x in y_pred]
    print('y train predict:',ypred_train)
    y_train = [list(x).index(max(x)) for x in y_train]
    print('y train actually:',y_train)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    plt.plot(range(1, epochs + 1), history.acc)
    plt.xlabel('Epochs')
    plt.ylabel(sheet+'Accuracy with Dropout'+str(num))
    plt.show()

    plt.plot(range(1, epochs + 1), history.loss)
    plt.xlabel('Epochs')
    plt.ylabel(sheet+'Loss with Dropout'+str(num))
    plt.show()
    model_name=sheet+"_"+str(num)+"epochs"+str(epochs)+".h5"
    # # 保存模型
    model.save("model/"+model_name)
    # # 加载整个模型
    # model.load_model('my_model.h5')
    return history.acc,history.loss,history.val_acc,history.val_loss
