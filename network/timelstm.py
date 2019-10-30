import keras
from keras.layers import Dense, Flatten, LSTM, MaxPooling1D, Convolution1D, Conv1D, \
    Dropout,TimeDistributed
from keras.models import Sequential
import matplotlib.pylab as plt
from openpyxl import load_workbook

from textcnn import AccuracyHistory,  readvec, reverse
import numpy as np

def toTimeDistribute(vec,y,l):
    vec=reverse(vec)
    y =reverse(y)
    lala=[]
    vector=[]
    Y=[]
    for i in range(len(vec)-l):
        for j in range(l):
            lala.append(vec[i+j])
        vector.append(lala)
        Y.append(y[i+l])
        lala=[]
    return vector,Y


def preprocess(num,Timelen,num_classes,type=1):
    wb1 = load_workbook('label.xlsx')
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    y=[]
    vec=[]

    if num ==4:
        for i in range(2274):
            # #row =5 三分类
            if sheet1.cell(i+1,4).value==-1:
                y.append(0)
            elif sheet1.cell(i+1,4).value==0:
                y.append(1)
            else:
                y.append(2)
    else:
        for i in range(2274):
        #二分类
            if sheet1.cell(i + 1, num).value == -1:
                y.append(0)
            else:
                y.append(1)

    if type==1:
        for i in range(2274):
            vec.append(sheet1.cell(i+1,5).value)
    else:
        vec=y

    vec, y = toTimeDistribute(vec, y, Timelen)
    # import matplotlib.pyplot as plt
    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(vec, y, test_size=0.4,shuffle=False)
    x_train=np.array(x_train)
    x_test=np.array(x_test)
    x_train = x_train.reshape(int(x_train.shape[0]), Timelen,1)
    x_test = x_test.reshape(int(x_test.shape[0]), Timelen,1)
    print('x_train shape:', x_train.shape)
    print(x_train.shape[0], 'train samples')
    print(x_test.shape[0], 'test samples')

    # convert class vectors to binary class matrices - this is for use in the
    # categorical_crossentropy loss below

    y_train = keras.utils.to_categorical(y_train, num_classes)
    y_test = keras.utils.to_categorical(y_test, num_classes)
    return x_train, y_train, x_test, y_test


def process(num=2,batch_size=128, \
    epochs = 300, Timelen = 5):
    if num == 4:
        num_classes = 3
    else:
        num_classes = 2


    x_train, y_train,x_test, y_test=preprocess(num,Timelen,num_classes)
    history = AccuracyHistory()
    model=Sequential()
    model.add(LSTM(Timelen))
    model.add(Dropout(0.4))
    model.add(Dense(num_classes, activation='softmax'))
    model.compile(loss=keras.losses.categorical_crossentropy,
                  optimizer=keras.optimizers.Adam(),
                  metrics=['accuracy'])
    model.fit(x_train, y_train,
              batch_size=batch_size,
              epochs=epochs,
              verbose=1,
              validation_data=(x_test, y_test),
              callbacks=[history])
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    plt.plot(range(1, epochs + 1), history.acc)
    plt.xlabel('Epochs')
    plt.ylabel(str(Timelen)+' Accuracy with Dropout'+str(num))
    plt.show()

    plt.plot(range(1, epochs + 1), history.loss)
    plt.xlabel('Epochs')
    plt.ylabel(str(Timelen)+' Loss with Dropout'+str(num))
    plt.show()
    model_name="only_price"+str(num)+"epochs"+str(epochs)+"_finaltime"+str(Timelen)+".h5"
    # # 保存模型
    model.save("model365/"+model_name)
    # # 加载整个模型
    # model.load_model('my_model.h5')
    return history.acc,history.loss,history.val_acc,history.val_loss

