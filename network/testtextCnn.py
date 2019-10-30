from keras import Input, Model
from keras.layers import Conv1D, Embedding, MaxPooling1D, Flatten, concatenate, Dropout, Dense, Conv2D, MaxPooling2D
from openpyxl import load_workbook
import numpy as np

def readvec(path,sheet):
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet)
    wb1 = load_workbook('b.xlsx')
    sheet1 = wb1.get_sheet_by_name('Sheet1')
    lala=[]
    vec=[]
    vector=[]
    y=[]
    date =sheet.cell(1,1).value

    for i in range(25589):
        new_date =sheet.cell(i+1,1).value
        if new_date!=date:
            print(date,new_date)
            date=new_date
            vector.append(vec)
            vec=[]
            lala=[]
        # print(date,new_date)
        for j in range(200):
            lala.append(float(sheet.cell(i+1,j+2).value))
        vec.append(lala)
        lala=[]
    # vector.append(vec)
    for i in range(2274):
        y.append(sheet1.cell(i+1,5).value)
        #y.append(int(sheet1.cell(i+1,8).value))
        # print(sheet1.cell(i+1,4).value)
    return vector,y

vec ,y =readvec('cleanedNewsUn200.xlsx','Sheet1' )
X=[]
lala=[]
for j in range(200): lala.append(0)
for i in range(len(vec)):
     while len(vec[i])<35:
         vec[i].append(lala)
     X.append(np.array(vec[i],dtype='float16').flatten())

X=np.array(X)
# print(X.shape)
y =np.array(y)
from sklearn.preprocessing import StandardScaler
X_scale = StandardScaler()
X = X_scale.fit_transform(X)
from sklearn.model_selection import train_test_split
x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=0.4)




def textcnn(maxlen=35,max_features=200,embed_size =200):
    comment_seq=Input(shape=[maxlen],name='x_seq')
    # print("not embed_size")
    # emb_comment =Embedding(35,200)(comment_seq)
    print (type(comment_seq))
    convs=[]
    filter_size=[5,7]
    for fsz in filter_size:
        l_conv=Conv1D(filters=100,kernel_size=fsz,activation='tanh')(comment_seq)# emb_comment)
        l_pool =MaxPooling1D(maxlen-fsz+1)(l_conv)
        l_pool =Flatten()(l_pool)
        convs.append(l_pool)
    merge =concatenate(convs,axis=1)

    out =Dropout(0.5)(merge)
    output =Dense(32,activation='relu')(out)
    output =Dense(units=1,activation='sigmoid')(out)

    model =Model([comment_seq],output)
    model.compile(loss="binary_crossentropy",optimizer="adam",metrics=['accuracy'])
    return model

maxlen=7000
max_feature=200
embed_size=200
model =textcnn(maxlen,max_feature,embed_size)
batch_size =64
epochs=10
model.fit(x_train,y_train,
          batch_size=batch_size,
          epochs=epochs,
          shuffle=True)

scores= model.evaluate(x_test,y_test)
print("test_loss:%f,accuracy %f"%(scores[0],scores[1]))

scores= model.evaluate(x_train,y_train)
print("train_loss:%f,accuracy %f"%(scores[0],scores[1]))
