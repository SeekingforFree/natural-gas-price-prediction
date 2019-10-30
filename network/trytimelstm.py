from openpyxl import Workbook

import timelstm

numlist=[1,2,3,4]
timelist=[30]
wb=Workbook()
sheet1 =wb.active
sheet1.title='acc_and_loss'
raw =1
for time in timelist:
    for num in numlist:
        acc,loss,testacc,testloss=timelstm.process(num=num,Timelen=time,epochs=300)

        sheet1.cell(raw,1).value = "Time:"+str(time)+" Num:"+str(num)
        raw+= 1
        for i in range(len(testacc)):
            sheet1.cell(raw,i+1).value =testacc[i]
        raw+=1
        for i in range(len(testloss)):
            sheet1.cell(raw,i+1).value =testloss[i]
        raw+=1
        for i in range(len(acc)):
            sheet1.cell(raw,i+1).value =acc[i]
        raw+=1
        for i in range(len(loss)):
            sheet1.cell(raw,i+1).value =loss[i]
        raw+=1
# wb.save("model365/only_time_acc_and_loss2_365.xlsx")