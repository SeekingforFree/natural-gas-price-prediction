from openpyxl import load_workbook
import matplotlib.pylab as plt
import numpy as np
wb = load_workbook('finalweek_acc_and_loss_nonsenti.xlsx')
sheet1 = wb.get_sheet_by_name('acc_and_loss')
epochs=30

# def plotC(acc, loss, str1):
#     T =np.array(range(1, epochs + 1))
#
#     from scipy.interpolate import spline
#     xnew = np.linspace(T.min(), T.max(), 50000)
#     acc_power_smooth = spline(T, acc, xnew)
#     plt.plot(xnew, acc_power_smooth)
#     plt.xlabel('Epochs')
#     plt.ylabel(str1 + ' Accuracy with Dropout')
#     plt.show()
#
#     loss_power_smooth = spline(T,loss,xnew)
#     plt.plot(xnew, loss_power_smooth)
#     plt.xlabel('Epochs')
#     plt.ylabel(str1 + ' Loss with Dropout' )
#     plt.show()


def plotC(acc, loss, str1):
    plt.plot(range(1, epochs + 1), acc)
    plt.xlabel('Epochs')
    plt.ylabel(str1 + ' Accuracy with Dropout')
    plt.show()

    plt.plot(range(1, epochs + 1), loss)
    plt.xlabel('Epochs')
    plt.ylabel(str1 + ' Loss with Dropout' )
    plt.show()


for i in range(9,10):
    acc1 = []
    loss1 = []
    acc2 = []
    loss2 = []
    for j in range(30):
        acc1.append(sheet1.cell(i * 5 + 2, j + 1).value)
        loss1.append(sheet1.cell(i * 5 + 3, j + 1).value)
        acc2.append(sheet1.cell(i * 5 + 4, j + 1).value)
        loss2.append(sheet1.cell(i * 5 + 5, j + 1).value)
    k = 0
    if int(i / 4) == 0:
        k = 7
    elif int(i / 4) == 1:
        k = 30
    elif int(i / 4) == 2:
        k = 60
    plotC(acc1, loss1, str(k) + ".." + str(i % 4 + 1)+str(i))
    plotC(acc2, loss2, str(k) + ".." + str(i % 4 + 1) + " Val")
