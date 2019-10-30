import matplotlib.pylab as plt
from openpyxl import load_workbook

import numpy as np
epochs=500


def plotC(acc, loss, str1):
    plt.plot(range(1, epochs + 1), acc)
    plt.xlabel('Epochs')
    plt.ylabel(str1 + ' Accuracy with Dropout')
    plt.show()

    plt.plot(range(1, epochs + 1), loss)
    plt.xlabel('Epochs')
    plt.ylabel(str1 + ' Loss with Dropout' )
    plt.show()


wb1 = load_workbook('only_time_acc_and_loss2_1.xlsx')
sheet1 = wb1.get_sheet_by_name('acc_and_loss')
for i in range(12):
    acc1=[]
    loss1=[]
    acc2=[]
    loss2=[]
    for j in range(500):
        acc1.append(sheet1.cell(i*5+2,j+1).value)
        loss1.append(sheet1.cell(i * 5 + 3, j + 1).value)
        acc2.append(sheet1.cell(i * 5 + 4, j + 1).value)
        loss2.append(sheet1.cell(i * 5 + 5, j + 1).value)
    k=0
    if int(i/4)==0:
        k=7
    elif int(i/4)==1:
        k=30
    elif int(i/4)==2:
        k=60
    plotC(acc1,loss1,str(k)+".."+str(i%4+1)+str(i))
    plotC(acc2, loss2,str(k)+".."+str(i%4+1)+ " Val"+str(i))
#为什么 train的acc没有val_acc好，其中程序的val选取的是test_set